from rest_framework import viewsets, permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from django.utils import timezone
from datetime import date
from attendance.models import Attendance, AttendanceCorrectionRequest, AttendanceStatus, AttendanceWorkMode, AttendanceMethod, CorrectionStatus, ShiftReport, FestivalHoliday
from attendance.serializers import (
    AttendanceSerializer, WFHAttendanceScanSerializer,
    AttendanceCorrectionSerializer, ShiftReportSerializer, FestivalHolidaySerializer
)
from django.utils import timezone
from datetime import date
from attendance.models import Attendance, AttendanceCorrectionRequest, AttendanceStatus, AttendanceWorkMode, AttendanceMethod, CorrectionStatus, ShiftReport
from attendance.serializers import (
    AttendanceSerializer, WFHAttendanceScanSerializer,
    AttendanceCorrectionSerializer, ShiftReportSerializer
)
from attendance.services import AttendanceEngine
from accounts.permissions import IsHR, IsCEO, IsEmployee
from accounts.models import Role
from audit.services import AuditService
from notifications.models import NotificationType
from notifications.services import NotificationService


class WFHAttendanceView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not hasattr(request.user, 'employee_profile'):
            return Response({'error': 'User does not have an active employee profile.'}, status=status.HTTP_400_BAD_REQUEST)

        employee = request.user.employee_profile
        today = date.today()
        now = timezone.now()

        # Rule 12: WFH Attendance requires an APPROVED WFH request for today
        if not AttendanceEngine.check_wfh_approval(employee, today):
            return Response({'error': 'WFH Attendance requires an APPROVED WFH request for today. Please apply for WFH first.'}, status=status.HTTP_400_BAD_REQUEST)

        # Rule 11: Leave conflict check
        if AttendanceEngine.check_leave_conflict(employee, today):
            return Response({'error': 'You are on APPROVED LEAVE today. WFH check-in is not allowed.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = WFHAttendanceScanSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        latitude = serializer.validated_data['latitude']
        longitude = serializer.validated_data['longitude']
        device_id = serializer.validated_data.get('device_id', 'WFH-MOBILE-WEB')

        attendance = Attendance.objects.filter(employee=employee, date=today).first()

        if attendance:
            if not attendance.check_out:
                attendance.check_out = now
                attendance.working_hours = AttendanceEngine.calculate_working_hours(attendance.check_in, now)
                attendance.status = AttendanceEngine.calculate_final_status(attendance)
                attendance.save()

                AuditService.log_action(
                    actor=request.user,
                    action='WFH_CHECK_OUT',
                    target_model='Attendance',
                    target_id=str(attendance.id),
                    reason=f"WFH Check-Out recorded for {employee.full_name}",
                    request=request
                )
                return Response({
                    'message': f"WFH Check-Out recorded for {employee.full_name}",
                    'type': 'CHECK_OUT',
                    'attendance': AttendanceSerializer(attendance).data
                }, status=status.HTTP_200_OK)
            else:
                return Response({'message': 'WFH Attendance already completed for today.'}, status=status.HTTP_200_OK)

        attendance = Attendance.objects.create(
            employee=employee,
            date=today,
            check_in=now,
            status=AttendanceStatus.WFH,
            work_mode=AttendanceWorkMode.WFH,
            attendance_method=AttendanceMethod.WEB_PORTAL,
            location_verified=True,
            latitude=latitude,
            longitude=longitude,
            device_id=device_id,
            taken_by=request.user
        )

        AuditService.log_action(
            actor=request.user,
            action='WFH_CHECK_IN',
            target_model='Attendance',
            target_id=str(attendance.id),
            new_values={'employee': employee.employee_id, 'lat': latitude, 'long': longitude},
            reason=f"WFH Check-In recorded for {employee.full_name}",
            request=request
        )

        return Response({
            'message': 'WFH Check-In recorded successfully',
            'type': 'CHECK_IN',
            'attendance': AttendanceSerializer(attendance).data
        }, status=status.HTTP_201_CREATED)

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all().order_by('-date', '-check_in')
    serializer_class = AttendanceSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = Attendance.objects.all().order_by('-date', '-check_in')

        if user.role in [Role.CEO, Role.HR, Role.SYSTEM_ADMIN]:
            # Apply optional filters
            emp_id = self.request.query_params.get('employee')
            dept_id = self.request.query_params.get('department')
            status_param = self.request.query_params.get('status')
            date_param = self.request.query_params.get('date')

            if emp_id:
                queryset = queryset.filter(employee_id=emp_id)
            if dept_id:
                queryset = queryset.filter(employee__department_id=dept_id)
            if status_param:
                queryset = queryset.filter(status=status_param)
            if date_param:
                queryset = queryset.filter(date=date_param)

            return queryset

        # Employee only sees own attendance
        if hasattr(user, 'employee_profile'):
            return queryset.filter(employee=user.employee_profile)
        return Attendance.objects.none()

    @action(detail=False, methods=['get'], url_path='logs')
    def logs(self, request):
        return self.list(request)

    @action(detail=False, methods=['get'], url_path='today-summary')
    def today_summary(self, request):
        today = date.today()
        attendances = Attendance.objects.filter(date=today)

        present_count = attendances.filter(status__in=[AttendanceStatus.PRESENT, AttendanceStatus.LATE]).count()
        wfh_count = attendances.filter(status=AttendanceStatus.WFH).count()
        leave_count = attendances.filter(status=AttendanceStatus.LEAVE).count()
        late_count = attendances.filter(status=AttendanceStatus.LATE).count()
        return Response({
            'date': today,
            'total_recorded': attendances.count(),
            'present_count': present_count,
            'wfh_count': wfh_count,
            'leave_count': leave_count,
            'late_count': late_count,
        })

    @action(detail=False, methods=['get'], url_path='today_summary')
    def today_summary_alias(self, request):
        return self.today_summary(request)

    @action(detail=False, methods=['post'], url_path='late_action', permission_classes=[permissions.IsAuthenticated])
    def late_action(self, request):
        user = request.user
        if not hasattr(user, 'employee_profile'):
            return Response({'error': 'User does not have an active employee profile.'}, status=status.HTTP_400_BAD_REQUEST)
        
        employee = user.employee_profile
        today = date.today()
        now = timezone.now()
        action_type = request.data.get('action')

        if action_type not in ['CLOCK_IN', 'MARK_ABSENT']:
            return Response({'error': 'Invalid action type.'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if attendance already exists
        attendance = Attendance.objects.filter(employee=employee, date=today).first()
        if attendance:
            return Response({'error': 'Attendance already marked for today.'}, status=status.HTTP_400_BAD_REQUEST)

        if action_type == 'MARK_ABSENT':
            attendance = Attendance.objects.create(
                employee=employee,
                date=today,
                check_in=now,
                check_out=now,
                working_hours=0,
                status=AttendanceStatus.ABSENT,
                work_mode=AttendanceWorkMode.OFFICE,
                attendance_method=AttendanceMethod.WEB_PORTAL,
                taken_by=user
            )
            message = "You have been marked as ABSENT for today."
            
        elif action_type == 'CLOCK_IN':
            is_wfh = AttendanceEngine.check_wfh_approval(employee, today)
            work_mode = AttendanceWorkMode.WFH if is_wfh else AttendanceWorkMode.OFFICE
            # Force status to PRESENT (since they just clicked clock in)
            attendance = Attendance.objects.create(
                employee=employee,
                date=today,
                check_in=now,
                status=AttendanceStatus.PRESENT,
                work_mode=work_mode,
                attendance_method=AttendanceMethod.WEB_PORTAL,
                taken_by=user
            )
            message = "You have successfully clocked in."

        # Find the corresponding notification and mark it as read
        try:
            from notifications.models import Notification, NotificationType
            notif = Notification.objects.filter(
                recipient=user,
                notification_type=NotificationType.LATE_CLOCK_IN_ALERT,
                is_read=False
            ).first()
            if notif:
                notif.is_read = True
                notif.save()
        except Exception:
            pass

        return Response({
            'message': message,
            'attendance': AttendanceSerializer(attendance).data
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='clock-in', permission_classes=[permissions.IsAuthenticated])
    def clock_in(self, request):
        user = request.user
        if not hasattr(user, 'employee_profile'):
            return Response({'error': 'User does not have an active employee profile.'}, status=status.HTTP_400_BAD_REQUEST)
        
        employee = user.employee_profile
        today = date.today()
        now = timezone.now()

        if AttendanceEngine.check_leave_conflict(employee, today):
            return Response({'error': 'You are on APPROVED LEAVE today. Clock-in is not allowed.'}, status=status.HTTP_400_BAD_REQUEST)

        is_wfh = AttendanceEngine.check_wfh_approval(employee, today)
        work_mode = AttendanceWorkMode.WFH if is_wfh else AttendanceWorkMode.OFFICE

        calc_status = AttendanceEngine.calculate_status(now, work_mode)

        attendance = Attendance.objects.filter(employee=employee, date=today).first()
        if attendance:
            # Check-in is strictly allowed once per day
            if attendance.check_in:
                return Response(
                    {'error': 'You have already checked in for today. Check-in is allowed only once per working day.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                # Update existing placeholder/un-marked record with check-in
                attendance.check_in = now
                attendance.status = calc_status
                attendance.work_mode = work_mode
                attendance.attendance_method = AttendanceMethod.WEB_PORTAL
                attendance.taken_by = user
                attendance.save()

                AuditService.log_action(
                    actor=user,
                    action='CLOCK_IN',
                    target_model='Attendance',
                    target_id=str(attendance.id),
                    reason=f"Clocked In via Web Portal ({work_mode})",
                    request=request
                )

                return Response({
                    'message': 'Clock-in successful',
                    'attendance': AttendanceSerializer(attendance).data
                }, status=status.HTTP_200_OK)

        attendance = Attendance.objects.create(
            employee=employee,
            date=today,
            check_in=now,
            status=calc_status,
            work_mode=work_mode,
            attendance_method=AttendanceMethod.WEB_PORTAL,
            location_verified=False,
            device_id='WEB-PORTAL-CLIENT',
            taken_by=user
        )

        AuditService.log_action(
            actor=user,
            action='CLOCK_IN',
            target_model='Attendance',
            target_id=str(attendance.id),
            reason=f"Clocked In via Web Portal ({work_mode})",
            request=request
        )

        return Response({
            'message': 'Clock-in successful',
            'attendance': AttendanceSerializer(attendance).data
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='clock_in', permission_classes=[permissions.IsAuthenticated])
    def clock_in_alias(self, request):
        return self.clock_in(request)

    @action(detail=False, methods=['post'], url_path='clock-out', permission_classes=[permissions.IsAuthenticated])
    def clock_out(self, request):
        user = request.user
        if not hasattr(user, 'employee_profile'):
            return Response({'error': 'User does not have an active employee profile.'}, status=status.HTTP_400_BAD_REQUEST)
        
        employee = user.employee_profile
        today = date.today()
        now = timezone.now()

        attendance = Attendance.objects.filter(employee=employee, date=today).first()
        if not attendance or not attendance.check_in:
            return Response({'error': 'No active checked-in session found for today. You must check in before checking out.'}, status=status.HTTP_400_BAD_REQUEST)

        # Checkout is strictly allowed once per day
        if attendance.check_out is not None:
            return Response({'error': 'You have already checked out for today. Checkout is allowed only once per working day.'}, status=status.HTTP_400_BAD_REQUEST)

        attendance.check_out = now
        attendance.working_hours = AttendanceEngine.calculate_working_hours(attendance.check_in, now)
        attendance.status = AttendanceEngine.calculate_final_status(attendance)
        attendance.save()

        AuditService.log_action(
            actor=user,
            action='CLOCK_OUT',
            target_model='Attendance',
            target_id=str(attendance.id),
            reason=f"Clocked Out via Web Portal. Working Hours: {attendance.working_hours}",
            request=request
        )

        return Response({
            'message': 'Clock-out successful',
            'attendance': AttendanceSerializer(attendance).data
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='clock_out', permission_classes=[permissions.IsAuthenticated])
    def clock_out_alias(self, request):
        return self.clock_out(request)

class AttendanceCorrectionViewSet(viewsets.ModelViewSet):
    queryset = AttendanceCorrectionRequest.objects.all().order_by('-created_at')
    serializer_class = AttendanceCorrectionSerializer

    def get_permissions(self):
        if self.action in ['approve', 'reject']:
            return [IsHR()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        if user.role in [Role.CEO, Role.HR, Role.SYSTEM_ADMIN]:
            qs = AttendanceCorrectionRequest.objects.all().order_by('-created_at')
        elif hasattr(user, 'employee_profile'):
            qs = AttendanceCorrectionRequest.objects.filter(employee=user.employee_profile).order_by('-created_at')
        else:
            return AttendanceCorrectionRequest.objects.none()

        status_param = self.request.query_params.get('status')
        if status_param:
            qs = qs.filter(status=status_param.upper())

        return qs

    def perform_create(self, serializer):
        employee = self.request.user.employee_profile
        existing_att = Attendance.objects.filter(employee=employee, date=serializer.validated_data['date']).first()
        instance = serializer.save(
            employee=employee,
            attendance=existing_att,
            original_check_in=existing_att.check_in if existing_att else None,
            original_check_out=existing_att.check_out if existing_att else None,
        )

        AuditService.log_action(
            actor=self.request.user,
            action='SUBMIT_ATTENDANCE_CORRECTION',
            target_model='AttendanceCorrectionRequest',
            target_id=str(instance.id),
            reason=f"Submitted correction request for date {serializer.validated_data['date']}",
            request=self.request
        )

        NotificationService.notify_management(
            title="Attendance Correction Requested",
            message=f"{employee.full_name} ({employee.employee_id}) submitted an attendance correction request for {instance.date}. Reason: {instance.reason}",
            notification_type=NotificationType.CORRECTION_SUBMITTED
        )

    @action(detail=True, methods=['post'], permission_classes=[IsHR])
    def approve(self, request, pk=None):
        correction = self.get_object()
        if correction.status != CorrectionStatus.PENDING:
            return Response({'error': 'Correction request is not pending'}, status=status.HTTP_400_BAD_REQUEST)

        correction.status = CorrectionStatus.APPROVED
        correction.reviewed_by = request.user
        correction.save()

        # Update or create attendance record without deleting original data
        attendance, created = Attendance.objects.get_or_create(
            employee=correction.employee,
            date=correction.date,
            defaults={
                'check_in': correction.requested_check_in,
                'check_out': correction.requested_check_out,
                'status': AttendanceStatus.PRESENT,
                'work_mode': AttendanceWorkMode.OFFICE,
                'attendance_method': AttendanceMethod.MANUAL_CORRECTION,
                'taken_by': request.user
            }
        )

        attendance.check_in = correction.requested_check_in
        attendance.check_out = correction.requested_check_out
        attendance.working_hours = AttendanceEngine.calculate_working_hours(correction.requested_check_in, correction.requested_check_out)
        attendance.status = AttendanceEngine.calculate_final_status(attendance)
        attendance.attendance_method = AttendanceMethod.MANUAL_CORRECTION
        attendance.save()

        # Notify Employee
        NotificationService.create_notification(
            recipient=correction.employee.user,
            title="Attendance Correction Approved",
            message=f"Your attendance correction request for {correction.date} has been APPROVED by HR/CEO.",
            notification_type='CORRECTION_APPROVED'
        )

        AuditService.log_action(
            actor=request.user,
            action='APPROVE_ATTENDANCE_CORRECTION',
            target_model='AttendanceCorrectionRequest',
            target_id=str(correction.id),
            old_values={'original_check_in': str(correction.original_check_in)},
            new_values={'requested_check_in': str(correction.requested_check_in)},
            reason=f"Approved attendance correction for {correction.employee.full_name}",
            request=request
        )

        return Response({'message': 'Attendance correction APPROVED and attendance updated.'})

    @action(detail=True, methods=['post'], permission_classes=[IsHR])
    def reject(self, request, pk=None):
        correction = self.get_object()
        reason = request.data.get('rejection_reason', 'Request rejected by authority.')
        correction.status = CorrectionStatus.REJECTED
        correction.reviewed_by = request.user
        correction.rejection_reason = reason
        correction.save()

        NotificationService.create_notification(
            recipient=correction.employee.user,
            title="Attendance Correction Rejected",
            message=f"Your attendance correction request for {correction.date} was REJECTED: {reason}",
            notification_type='CORRECTION_REJECTED'
        )

        AuditService.log_action(
            actor=request.user,
            action='REJECT_ATTENDANCE_CORRECTION',
            target_model='AttendanceCorrectionRequest',
            target_id=str(correction.id),
            reason=f"Rejected attendance correction for {correction.employee.full_name}: {reason}",
            request=request
        )

        return Response({'message': 'Attendance correction REJECTED.'})

class ShiftReportViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ShiftReportSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = ShiftReport.objects.all().select_related('employee', 'employee__department').order_by('-date', '-created_at')

        # Restrict standard employees to their own shift reports logs
        if user.role not in [Role.CEO, Role.HR]:
            if hasattr(user, 'employee_profile'):
                queryset = queryset.filter(employee=user.employee_profile)
            else:
                return ShiftReport.objects.none()

        # Query Filters for CEO / HR / Employee
        date_param = self.request.query_params.get('date')
        if date_param:
            queryset = queryset.filter(date=date_param)

        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])
        elif start_date:
            queryset = queryset.filter(date__gte=start_date)
        elif end_date:
            queryset = queryset.filter(date__lte=end_date)

        employee_param = self.request.query_params.get('employee')
        if employee_param:
            queryset = queryset.filter(employee_id=employee_param)

        emp_code = self.request.query_params.get('employee_id')
        if emp_code:
            queryset = queryset.filter(employee__employee_id__iexact=emp_code)

        dept_param = self.request.query_params.get('department')
        if dept_param:
            queryset = queryset.filter(employee__department_id=dept_param)

        search_param = self.request.query_params.get('search')
        if search_param:
            queryset = queryset.filter(
                models.Q(employee__full_name__icontains=search_param) |
                models.Q(employee__employee_id__icontains=search_param) |
                models.Q(report_content__icontains=search_param)
            )

        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        emp_id = self.request.data.get('employee')
        if emp_id and user.role in [Role.CEO, Role.HR]:
            from employees.models import Employee
            target_emp = Employee.objects.filter(id=emp_id).first()
            if target_emp:
                serializer.save(employee=target_emp)
                return
        serializer.save(employee=user.employee_profile)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        import io
        from django.http import HttpResponse
        from attendance.models import Attendance
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        reports = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Shift Reports"
        ws.views.sheetView[0].showGridLines = True

        # Styles definition
        title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')

        subtitle_font = Font(name='Segoe UI', size=10, italic=True, color='CBD5E1')
        subtitle_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

        kpi_val_font = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
        kpi_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

        header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

        data_font = Font(name='Segoe UI', size=10, color='1E293B')
        data_font_bold = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
        mono_font = Font(name='Consolas', size=10, color='334155')

        thin_side = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

        zebra_fill_a = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        zebra_fill_b = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        # 1. Main Title Banner (Row 1 & 2)
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = "ENTERPRISE DAILY SHIFT REPORTS"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36

        ws.merge_cells('A2:K2')
        sub_cell = ws['A2']
        sub_cell.value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} (UTC) | Confidential Corporate Record | Active Record Count: {reports.count()} Reports"
        sub_cell.font = subtitle_font
        sub_cell.fill = subtitle_fill
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        # 2. Executive KPI Summary Cards (Row 4)
        total_reports_cnt = reports.count()

        kpis = [
            ('A4:D4', 'TOTAL SHIFT REPORTS', f"{total_reports_cnt} Logged"),
        ]

        for cell_range, title, val in kpis:
            ws.merge_cells(cell_range)
            start_cell_addr = cell_range.split(':')[0]
            top_cell = ws[start_cell_addr]
            top_cell.value = f"{title}: {val}"
            top_cell.font = kpi_val_font
            top_cell.fill = kpi_fill
            top_cell.alignment = Alignment(horizontal='center', vertical='center')
            top_cell.border = cell_border

        ws.row_dimensions[4].height = 28

        # 3. Table Headers (Row 6)
        headers = [
            ("Date", 15, 'center'),
            ("Employee ID", 16, 'center'),
            ("Employee Name", 26, 'left'),
            ("Corporate Email", 28, 'left'),
            ("Department", 22, 'left'),
            ("Check-In Time", 16, 'center'),
            ("Check-Out Time", 16, 'center'),
            ("Shift Hours", 16, 'center'),
            ("Attendance Status", 18, 'center'),
            ("Work Report", 80, 'left'),
            ("Logged Timestamp", 22, 'center')
        ]

        ws.row_dimensions[6].height = 30
        for col_idx, (header_text, width, align) in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col_idx)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # 4. Data Rows (Row 7+)
        current_row = 7
        for idx, report in enumerate(reports):
            att = Attendance.objects.filter(employee=report.employee, date=report.date).first()
            check_in_str = att.check_in.strftime('%H:%M:%S') if (att and att.check_in) else '--'
            check_out_str = att.check_out.strftime('%H:%M:%S') if (att and att.check_out) else '--'
            shift_hours = float(att.working_hours) if (att and att.working_hours) else 0.00
            att_status = att.status if att else 'NOT_MARKED'

            row_fill = zebra_fill_a if idx % 2 == 0 else zebra_fill_b
            ws.row_dimensions[current_row].height = 50 # Make rows a bit taller for reports

            row_values = [
                (report.date.strftime('%Y-%m-%d'), mono_font, 'center'),
                (report.employee.employee_id, mono_font, 'center'),
                (report.employee.full_name, data_font_bold, 'left'),
                (report.employee.email, data_font, 'left'),
                (report.employee.department.name if report.employee.department else 'Unassigned', data_font, 'left'),
                (check_in_str, mono_font, 'center'),
                (check_out_str, mono_font, 'center'),
                (f"{shift_hours:.2f} hrs", mono_font, 'center'),
                (att_status, data_font, 'center'),
                (report.report_content or '', data_font, 'left'),
                (report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else '', mono_font, 'center')
            ]

            for col_idx, (val, font, align) in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = font
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
                cell.border = cell_border

            current_row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Daily_Shift_Report_{timezone.localdate().strftime('%Y_%m_%d')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        from attendance.models import Attendance

        reports = self.get_queryset()
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"daily_shift_report_{timezone.localdate().strftime('%Y_%m_%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # UTF-8 BOM so Excel opens CSV cleanly without encoding issues
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow([
            'Shift Date',
            'Employee ID',
            'Employee Name',
            'Corporate Email',
            'Department',
            'Check-In Time',
            'Check-Out Time',
            'Shift Hours Worked',
            'Attendance Status',
            'Work Report',
            'Logged At'
        ])

        for report in reports:
            att = Attendance.objects.filter(employee=report.employee, date=report.date).first()
            check_in_str = att.check_in.strftime('%H:%M:%S') if (att and att.check_in) else 'N/A'
            check_out_str = att.check_out.strftime('%H:%M:%S') if (att and att.check_out) else 'N/A'
            total_hours = f"{float(att.working_hours):.2f} hrs" if (att and att.working_hours) else "0.00 hrs"
            att_status = att.status if att else 'NOT_MARKED'

            writer.writerow([
                report.date.strftime('%Y-%m-%d'),
                report.employee.employee_id,
                report.employee.full_name,
                report.employee.email,
                report.employee.department.name if report.employee.department else 'Unassigned',
                check_in_str,
                check_out_str,
                total_hours,
                att_status,
                report.report_content or '',
                report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else ''
            ])

        return response

class FestivalHolidayViewSet(viewsets.ModelViewSet):
    queryset = FestivalHoliday.objects.all().order_by('date')
    serializer_class = FestivalHolidaySerializer

    def get_permissions(self):
        ws = wb.active
        ws.title = "Daily Shift Reports"
        ws.views.sheetView[0].showGridLines = True

        # Styles definition
        title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')

        subtitle_font = Font(name='Segoe UI', size=10, italic=True, color='CBD5E1')
        subtitle_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

        kpi_val_font = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
        kpi_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

        header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

        data_font = Font(name='Segoe UI', size=10, color='1E293B')
        data_font_bold = Font(name='Segoe UI', size=10, bold=True, color='0F172A')
        mono_font = Font(name='Consolas', size=10, color='334155')

        thin_side = Side(border_style="thin", color="E2E8F0")
        cell_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

        zebra_fill_a = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        zebra_fill_b = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        # 1. Main Title Banner (Row 1 & 2)
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = "ENTERPRISE DAILY SHIFT REPORTS"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36

        ws.merge_cells('A2:K2')
        sub_cell = ws['A2']
        sub_cell.value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} (UTC) | Confidential Corporate Record | Active Record Count: {reports.count()} Reports"
        sub_cell.font = subtitle_font
        sub_cell.fill = subtitle_fill
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        # 2. Executive KPI Summary Cards (Row 4)
        total_reports_cnt = reports.count()

        kpis = [
            ('A4:D4', 'TOTAL SHIFT REPORTS', f"{total_reports_cnt} Logged"),
        ]

        for cell_range, title, val in kpis:
            ws.merge_cells(cell_range)
            start_cell_addr = cell_range.split(':')[0]
            top_cell = ws[start_cell_addr]
            top_cell.value = f"{title}: {val}"
            top_cell.font = kpi_val_font
            top_cell.fill = kpi_fill
            top_cell.alignment = Alignment(horizontal='center', vertical='center')
            top_cell.border = cell_border

        ws.row_dimensions[4].height = 28

        # 3. Table Headers (Row 6)
        headers = [
            ("Date", 15, 'center'),
            ("Employee ID", 16, 'center'),
            ("Employee Name", 26, 'left'),
            ("Corporate Email", 28, 'left'),
            ("Department", 22, 'left'),
            ("Check-In Time", 16, 'center'),
            ("Check-Out Time", 16, 'center'),
            ("Shift Hours", 16, 'center'),
            ("Attendance Status", 18, 'center'),
            ("Work Report", 80, 'left'),
            ("Logged Timestamp", 22, 'center')
        ]

        ws.row_dimensions[6].height = 30
        for col_idx, (header_text, width, align) in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col_idx)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # 4. Data Rows (Row 7+)
        current_row = 7
        for idx, report in enumerate(reports):
            att = Attendance.objects.filter(employee=report.employee, date=report.date).first()
            check_in_str = att.check_in.strftime('%H:%M:%S') if (att and att.check_in) else '--'
            check_out_str = att.check_out.strftime('%H:%M:%S') if (att and att.check_out) else '--'
            shift_hours = float(att.working_hours) if (att and att.working_hours) else 0.00
            att_status = att.status if att else 'NOT_MARKED'

            row_fill = zebra_fill_a if idx % 2 == 0 else zebra_fill_b
            ws.row_dimensions[current_row].height = 50 # Make rows a bit taller for reports

            row_values = [
                (report.date.strftime('%Y-%m-%d'), mono_font, 'center'),
                (report.employee.employee_id, mono_font, 'center'),
                (report.employee.full_name, data_font_bold, 'left'),
                (report.employee.email, data_font, 'left'),
                (report.employee.department.name if report.employee.department else 'Unassigned', data_font, 'left'),
                (check_in_str, mono_font, 'center'),
                (check_out_str, mono_font, 'center'),
                (f"{shift_hours:.2f} hrs", mono_font, 'center'),
                (att_status, data_font, 'center'),
                (report.report_content or '', data_font, 'left'),
                (report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else '', mono_font, 'center')
            ]

            for col_idx, (val, font, align) in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                cell.font = font
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
                cell.border = cell_border

            current_row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Daily_Shift_Report_{timezone.localdate().strftime('%Y_%m_%d')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        import csv
        from django.http import HttpResponse
        from attendance.models import Attendance

        reports = self.get_queryset()
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"daily_shift_report_{timezone.localdate().strftime('%Y_%m_%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # UTF-8 BOM so Excel opens CSV cleanly without encoding issues
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow([
            'Shift Date',
            'Employee ID',
            'Employee Name',
            'Corporate Email',
            'Department',
            'Check-In Time',
            'Check-Out Time',
            'Shift Hours Worked',
            'Attendance Status',
            'Work Report',
            'Logged At'
        ])

        for report in reports:
            att = Attendance.objects.filter(employee=report.employee, date=report.date).first()
            check_in_str = att.check_in.strftime('%H:%M:%S') if (att and att.check_in) else 'N/A'
            check_out_str = att.check_out.strftime('%H:%M:%S') if (att and att.check_out) else 'N/A'
            total_hours = f"{float(att.working_hours):.2f} hrs" if (att and att.working_hours) else "0.00 hrs"
            att_status = att.status if att else 'NOT_MARKED'

            writer.writerow([
                report.date.strftime('%Y-%m-%d'),
                report.employee.employee_id,
                report.employee.full_name,
                report.employee.email,
                report.employee.department.name if report.employee.department else 'Unassigned',
                check_in_str,
                check_out_str,
                total_hours,
                att_status,
                report.report_content or '',
                report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else ''
            ])

        return response

class FestivalHolidayViewSet(viewsets.ModelViewSet):
    queryset = FestivalHoliday.objects.all().order_by('date')
    serializer_class = FestivalHolidaySerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsHR()]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='calendar-events')
    def calendar_events(self, request):
        year = request.query_params.get('year', date.today().year)
        month = request.query_params.get('month', date.today().month)

        try:
            year = int(year)
            month = int(month)
        except ValueError:
            return Response({'error': 'Invalid year or month'}, status=status.HTTP_400_BAD_REQUEST)

        user = request.user
        events = []

        from core.models import Holiday
        
        # 1. Get Enterprise Settings Holidays
        holidays = Holiday.objects.filter(date__year=year, date__month=month)
        for h in holidays:
            events.append({
                'id': f'hol_{h.id}',
                'title': h.title,
                'date': h.date,
                'type': 'FESTIVAL',
                'festival_type': 'OPTIONAL' if h.is_optional else 'GENERAL',
                'color': '#f97316' if h.is_optional else '#ef4444',
                'allDay': True
            })

        from leaves.models import LeaveRequest, LeaveStatus
        from wfh.models import WFHRequest
        from datetime import timedelta
        from django.db.models import Q

        # Filters
        filter_emp_id = request.query_params.get('employee_id')
        filter_dept_id = request.query_params.get('department_id')
        filter_status = request.query_params.get('status') # 'APPROVED' | 'PENDING' | 'ALL'

        if user.role in [Role.CEO, Role.HR, Role.SYSTEM_ADMIN]:
            leave_qs = LeaveRequest.objects.filter(
                Q(start_date__year=year, start_date__month=month) |
                Q(end_date__year=year, end_date__month=month)
            ).select_related('employee', 'leave_type', 'employee__department', 'employee__user')

            wfh_qs = WFHRequest.objects.filter(date__year=year, date__month=month).select_related('employee')
            past_wfhs = Attendance.objects.filter(date__year=year, date__month=month, work_mode=AttendanceWorkMode.WFH).select_related('employee')

            if filter_emp_id:
                leave_qs = leave_qs.filter(employee_id=filter_emp_id)
                wfh_qs = wfh_qs.filter(employee_id=filter_emp_id)
                past_wfhs = past_wfhs.filter(employee_id=filter_emp_id)
            if filter_dept_id:
                leave_qs = leave_qs.filter(employee__department_id=filter_dept_id)
                wfh_qs = wfh_qs.filter(employee__department_id=filter_dept_id)
                past_wfhs = past_wfhs.filter(employee__department_id=filter_dept_id)
            if filter_status and filter_status != 'ALL':
                leave_qs = leave_qs.filter(status=filter_status)
            else:
                leave_qs = leave_qs.filter(status__in=[LeaveStatus.APPROVED, LeaveStatus.PENDING])
        else:
            if not hasattr(user, 'employee_profile'):
                return Response({'events': events})
            emp = user.employee_profile
            leave_qs = LeaveRequest.objects.filter(
                employee=emp,
                status__in=[LeaveStatus.APPROVED, LeaveStatus.PENDING]
            ).filter(
                Q(start_date__year=year, start_date__month=month) |
                Q(end_date__year=year, end_date__month=month)
            ).select_related('employee', 'leave_type', 'employee__department', 'employee__user')

            wfh_qs = WFHRequest.objects.filter(employee=emp, date__year=year, date__month=month).select_related('employee')
            past_wfhs = Attendance.objects.filter(employee=emp, date__year=year, date__month=month, work_mode=AttendanceWorkMode.WFH).select_related('employee')

        # Multi-day leave expansion across all calendar days
        for l in leave_qs:
            is_approved = (l.status == LeaveStatus.APPROVED)
            color = '#10b981' if is_approved else '#f59e0b' # Green for Approved, Amber for Planned/Pending
            status_text = 'Approved Leave' if is_approved else 'Planned (Pending)'

            curr = l.start_date
            while curr <= l.end_date:
                if curr.year == year and curr.month == month:
                    events.append({
                        'id': f'leave_{l.id}_{curr.isoformat()}',
                        'leave_id': l.id,
                        'title': f"{l.employee.full_name}: {l.leave_type.code or 'Leave'}" if user.role in [Role.CEO, Role.HR, Role.SYSTEM_ADMIN] else f"{l.leave_type.name} ({status_text})",
                        'date': curr.isoformat(),
                        'type': 'LEAVE',
                        'status': l.status,
                        'leave_status': l.status,
                        'leave_type_code': l.leave_type.code if l.leave_type else 'L',
                        'leave_type_name': l.leave_type.name if l.leave_type else 'Leave',
                        'employee_id': l.employee.id,
                        'employee_name': l.employee.full_name,
                        'department_name': l.employee.department.name if l.employee.department else 'Unassigned',
                        'avatar': l.employee.user.avatar.url if (l.employee.user and l.employee.user.avatar) else '',
                        'reason': l.reason,
                        'number_of_days': l.number_of_days,
                        'start_date': l.start_date.isoformat(),
                        'end_date': l.end_date.isoformat(),
                        'color': color,
                        'allDay': True
                    })
                curr += timedelta(days=1)

        for w in wfh_qs:
            events.append({
                'id': f'wfh_{w.id}',
                'title': f'WFH: {w.employee.full_name}' if user.role in [Role.CEO, Role.HR, Role.SYSTEM_ADMIN] else 'Approved WFH',
                'date': w.date.isoformat() if hasattr(w.date, 'isoformat') else str(w.date),
                'type': 'WFH_REQUEST',
                'status': w.status,
                'employee_name': w.employee.full_name,
                'color': '#3b82f6',
                'allDay': True
            })

        for a in past_wfhs:
            events.append({
                'id': f'att_wfh_{a.id}',
                'title': f'Worked WFH: {a.employee.full_name}' if user.role in [Role.CEO, Role.HR, Role.SYSTEM_ADMIN] else 'Worked From Home',
                'date': a.date.isoformat() if hasattr(a.date, 'isoformat') else str(a.date),
                'type': 'PAST_WFH',
                'employee_name': a.employee.full_name,
                'color': '#0ea5e9',
                'allDay': True
            })

        return Response({'events': events})
